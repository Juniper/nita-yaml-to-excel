#!/usr/bin/python3

""" ********************************************************

Project: nita-yaml-to-excel
Version: 21.7

Copyright (c) Juniper Networks, Inc., 2021. All rights reserved.

Notice and Disclaimer: This code is licensed to you under the Apache 2.0 License (the "License"). You may not use this code except in compliance with the License. This code is not an official Juniper product. You can obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0.html

SPDX-License-Identifier: Apache-2.0

Third-Party Code: This code may depend on other components under separate copyright notice and license terms. Your use of the source code for those components is subject to the terms and conditions of the respective license as noted in the Third-Party source code file.

********************************************************"""

import ast
import json
import logging
import os
import sys
from collections import OrderedDict

import yaml
from openpyxl import load_workbook

logging.basicConfig(stream=sys.stderr, level=logging.WARNING,
                    format='%(asctime)s: %(levelname)s: %(message)s')

pre_defined_unique_identifiers = ['id', 'name', 'group']


def ordered_dump(data, stream=None, Dumper=yaml.Dumper, **kwds):
    class OrderedDumper(Dumper):
        pass

    def _dict_representer(dumper, data):
        return dumper.represent_mapping(
            yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
            data.items())
    OrderedDumper.add_representer(OrderedDict, _dict_representer)
    OrderedDumper.add_representer(
        type(None),
        lambda dumper, value: dumper.represent_scalar(
            u'tag:yaml.org,2002:null', '')
    )
    return yaml.dump(data, stream, OrderedDumper, **kwds)


def stripper(self, data):
    new_data = OrderedDict()
    for k, v in data.items():
        if isinstance(v, OrderedDict):
            v = stripper(self, v)
        if not v in (list, OrderedDict()):
            new_data[k] = v
    return new_data


class ExcelToYaml(object):

    # -----------------------------------------------------------------------
    # CONSTRUCTOR
    # -----------------------------------------------------------------------

    def __init__(self, workbook_name_param, dest_dir_param):

        if workbook_name_param == "":
            logging.error("Workbook name required.")
            raise Exception("Workbook name required.")
        if dest_dir_param == "":
            logging.error("Destination directory required.")
            raise Exception("Destination directory required.")

        self.workbook_name = workbook_name_param
        self.dest_dir = dest_dir_param
        self.sheet_data = OrderedDict()

    def write_base_sheet_data(self, base_sheet_data, dest_dir):

        logging.debug("Base sheet data ::: %s ", base_sheet_data)
        for host_file in base_sheet_data:

            # 			temp_dict_data = ast.literal_eval(json.dumps(base_sheet_data[host_file]))
            # base_yaml_content=ordered_dump(base_sheet_data[host_file], Dumper=yaml.SafeD

            base_sheet_data[host_file] = stripper(
                self, base_sheet_data[host_file])

            base_yaml_content = ordered_dump(OrderedDict(
                base_sheet_data[host_file]), Dumper=yaml.SafeDumper, default_flow_style=False, explicit_start=True)
            logging.debug("Final data-")
            logging.debug(base_yaml_content)

            if not os.path.isdir(dest_dir):
                os.makedirs(dest_dir)

            if host_file:
                host_file_dir = dest_dir + '/' + os.path.dirname(host_file)
                if not os.path.isdir(host_file_dir):
                    os.makedirs(host_file_dir)
                with open(dest_dir + '/' + host_file, 'w') as outfile:
                    outfile.write(base_yaml_content)

    def map_key_value(self, key, value, param_dict):

        splitted_keys = key.split('.')
        mapped_dict = param_dict
        value = self.parse_cell_value(value)
        if len(splitted_keys) > 1:

            index = key.find(".")
            first_str = key[: index]
            sliced_str = key[index + 1:]

            if first_str in mapped_dict.keys():
                temp_recursive_dict = mapped_dict[first_str]
                temp_dict = self.map_key_value(
                    sliced_str, value, temp_recursive_dict)
                temp_recursive_dict.update(OrderedDict(temp_dict))
            else:
                temp_dict = self.map_key_value(
                    sliced_str, value, OrderedDict())
                mapped_dict.update(OrderedDict({first_str: temp_dict}))
        else:
            if value != "":
                mapped_dict = OrderedDict({key: value})

        return mapped_dict

    def process_data(self, key, value, hostname, sheet_name):

        logging.debug("Key :: %s Value :: %s ", key, value)

        if hostname in self.sheet_data.keys():
            temp_host_data = self.sheet_data[hostname]
            if sheet_name in temp_host_data.keys():
                temp_sheet_data = temp_host_data[sheet_name]
                temp_dict = self.map_key_value(key, value, temp_sheet_data)
                temp_sheet_data.update(temp_dict)
            else:
                logging.debug("New sheet data ::: ")
                temp_dict = self.map_key_value(key, value, OrderedDict())
                temp_sheet_data = OrderedDict()
                temp_sheet_data.update({sheet_name: temp_dict})
                temp_host_data.update(temp_sheet_data)

        else:
            logging.debug("New hostname found..")
            if hostname != "":
                temp_dict = self.map_key_value(key, value, OrderedDict())
                temp_sheet_data = OrderedDict()
                temp_sheet_data.update({sheet_name: temp_dict})
                host_dict = OrderedDict()
                host_dict.update({hostname: temp_sheet_data})
                self.sheet_data.update(host_dict)

    def get_appropriate_array_data(self, unique_field, unique_value, old_data):

        if old_data:
            if unique_field != None and unique_value != None:
                logging.debug("Got it.... %s ", type(old_data))
                if type(old_data) is list:
                    for old_dict_data in old_data:
                        if type(old_dict_data) is OrderedDict:
                            if unique_field in old_dict_data.keys():
                                if old_dict_data[unique_field] == unique_value:
                                    return old_dict_data

        return OrderedDict()

    def build_recursive_data(self, parentKey, dict_data, old_data):

        logging.debug("\ndict_data ::: %s ", dict_data)
        logging.debug(" >>> old_data :: %s ", old_data)

        unique_field = ""
        unique_identifier = ""
        first_level_dict_data = OrderedDict()
        hierarchy_dict_data = OrderedDict()

        existing_data = OrderedDict()

        for dict_key in dict_data:
            if old_data:
                if dict_key.find("@") >= 0:
                    logging.debug("Unique identifier ::: ")
                    unique_identifier = dict_key

                if dict_key.find("@") >= 0:
                    index = dict_key.find("@")
                    unique_field = dict_key[index + 1:]

                    logging.debug(
                        "Going to get old dict data .................. %s ", unique_field)
                    existing_data = self.get_appropriate_array_data(
                        unique_field, dict_data[unique_identifier], old_data)
                    if existing_data:
                        first_level_dict_data = existing_data

        for dict_key in dict_data:
            if dict_key.find("@") >= 0:
                unique_identifier = dict_key

            value = dict_data[dict_key]

            if old_data:
                if dict_key.find("@") >= 0:
                    index = dict_key.find("@")
                    unique_field = dict_key[index + 1:]

                    logging.debug(
                        "Going to get old dict data ..................")
                    existing_data = self.get_appropriate_array_data(
                        unique_field, dict_data[unique_identifier], old_data)
                    if existing_data:
                        first_level_dict_data = existing_data

            if dict_key.find("+") >= 0:
                temp_dic_data = None
                index = dict_key.find("+")
                temp_key = dict_key[: index]

                array_data = OrderedDict()
                previous_data = OrderedDict()
                temp_array_hierarchy_data = OrderedDict()
                if type(value) is OrderedDict or type(value) is list:
                    if old_data:
                        if type(old_data) is OrderedDict:
                            if temp_key in old_data.keys():
                                temp_array_hierarchy_data = old_data[temp_key]
                                first_level_dict_data.update(
                                    OrderedDict({temp_key: old_data[temp_key]}))

                        elif type(old_data) is list:
                            for old_data_row in old_data:
                                if temp_key in old_data_row.keys():
                                    if (parentKey == None and (unique_field in first_level_dict_data.keys() and unique_field in old_data_row.keys()
                                                               and old_data_row[unique_field] == first_level_dict_data[unique_field])):
                                        temp_array_hierarchy_data = old_data_row[temp_key]
                                        parentKey = None
                                        break
                                    if (parentKey in old_data_row.keys() and parentKey in first_level_dict_data.keys() and
                                            old_data_row[parentKey] == first_level_dict_data[parentKey]):
                                        temp_array_hierarchy_data = old_data_row[temp_key]
                                        parentKey = None
                                        break

                    temp_dic_data = self.build_recursive_data(
                        parentKey, value, temp_array_hierarchy_data)

                    if temp_dic_data:
                        if "new_data" in temp_dic_data.keys():
                            array_data = temp_dic_data["new_data"]
                        if "existing_data" in temp_dic_data.keys():
                            previous_data = temp_dic_data["existing_data"]
                else:
                    if old_data:
                        if type(old_data) is OrderedDict:
                            if temp_key in old_data.keys():
                                first_level_dict_data.update(
                                    OrderedDict({temp_key: old_data[temp_key]}))

                    array_data = value

                if isinstance(array_data, int) or array_data:
                    logging.debug("Not empty.....")
                    if temp_key in first_level_dict_data:
                        temp_array_data = first_level_dict_data[temp_key]
                        if not previous_data:
                            temp_array_data.append(array_data)
                    else:
                        logging.debug("Key not found")
                        first_level_dict_data.update(
                            OrderedDict({temp_key: [array_data]}))

            else:
                if type(value) is not OrderedDict and type(value) is not list:

                    logging.debug("Not a dict and list %s %s %s",
                                  dict_key, value, type(value))

                    unique_field = dict_key
                    if dict_key.find("@") >= 0:
                        index = dict_key.find("@")
                        unique_field = dict_key[index + 1:]

                    if dict_data[dict_key] != "":
                        first_level_dict_data.update(OrderedDict(
                            {unique_field: dict_data[dict_key]}))
                else:
                    logging.debug("Process hierarchy data %s ", dict_key)
                    if dict_data[dict_key]:
                        hierarchy_dict_data.update(
                            OrderedDict({dict_key: dict_data[dict_key]}))

        if hierarchy_dict_data:
            for hierarchy_key in hierarchy_dict_data:

                temp_array_hierarchy_data = first_level_dict_data
                if hierarchy_key in first_level_dict_data.keys():
                    temp_array_hierarchy_data = first_level_dict_data[hierarchy_key]

                if type(old_data) is OrderedDict:
                    if hierarchy_key in old_data.keys():
                        temp_array_hierarchy_data = old_data[hierarchy_key]

                temp_dic_data = self.build_recursive_data(
                    parentKey, hierarchy_dict_data[hierarchy_key], temp_array_hierarchy_data)

                temp_hierarchy_data = OrderedDict()
                if temp_dic_data:
                    if "new_data" in temp_dic_data.keys():
                        temp_hierarchy_data = temp_dic_data["new_data"]

                if type(old_data) is OrderedDict:
                    if hierarchy_key in old_data.keys():
                        temp_hierarchy_data.update(old_data[hierarchy_key])

                if temp_hierarchy_data:
                    first_level_dict_data.update(OrderedDict(
                        {hierarchy_key: temp_hierarchy_data}))

        logging.debug("First level dict: %s ", first_level_dict_data)

        return OrderedDict({"new_data":  first_level_dict_data, "existing_data": existing_data})

    def add_hierarical_data(self, parentKey, dict_data, hostname, sheet_name, is_list=True):

        if hostname in self.sheet_data.keys():
            temp_host_data = self.sheet_data[hostname]
            if sheet_name in temp_host_data.keys():
                temp_sheet_data = temp_host_data[sheet_name]
                structured_data = self.build_recursive_data(
                    parentKey, dict_data, temp_sheet_data)
                logging.debug("============================================")
                logging.debug(structured_data)
                if structured_data:
                    if "existing_data" in structured_data.keys():
                        if not structured_data["existing_data"]:
                            structured_data = structured_data["new_data"]
                            if is_list:
                                temp_sheet_data.append(structured_data)
                            else:
                                temp_sheet_data.update(structured_data)
                    else:
                        structured_data = structured_data["new_data"]
                        if is_list:
                            temp_sheet_data.append(structured_data)
                        else:
                            temp_sheet_data.update(structured_data)
            else:
                logging.debug("Going to build new sheet data...")
                structured_data = self.build_recursive_data(
                    parentKey, dict_data, OrderedDict())

                if structured_data:
                    if "new_data" in structured_data.keys():
                        structured_data = structured_data["new_data"]

                temp_sheet_data = OrderedDict()
                if is_list:
                    temp_dict = [structured_data]
                    temp_sheet_data.update(
                        OrderedDict({sheet_name: temp_dict}))
                else:
                    temp_sheet_data.update(OrderedDict(
                        {sheet_name: structured_data}))
                temp_host_data.update(temp_sheet_data)

        else:
            logging.debug("\nGoing to build new host data...")
            structured_data = self.build_recursive_data(
                parentKey, dict_data, OrderedDict())

            if structured_data:
                if "new_data" in structured_data.keys():
                    structured_data = structured_data["new_data"]

            temp_sheet_data = OrderedDict()
            if is_list:
                temp_dict = [structured_data]
                temp_sheet_data.update(OrderedDict({sheet_name: temp_dict}))
            else:
                temp_sheet_data.update(OrderedDict(
                    {sheet_name: structured_data}))
            host_dict = OrderedDict()
            host_dict.update(OrderedDict({hostname: temp_sheet_data}))
            self.sheet_data.update(host_dict)

    def process_list_data(self, parentKey, dict_data, hostname, sheet_name):

        logging.debug(dict_data)
        logging.debug("keys::::::::::::::::::::::::: %s ", dict_data.keys())

        is_nested = False
        for dict_key in dict_data:
            if dict_key.find("+") > 0:
                is_nested = True

        if is_nested:
            if sheet_name.find("+") > 0:
                sheet_index = sheet_name.find("+")
                sheet_name = sheet_name[: sheet_index]

                self.add_hierarical_data(
                    parentKey, dict_data, hostname, sheet_name, True)

            else:
                logging.debug("List within in dict...")

                self.add_hierarical_data(
                    parentKey, dict_data, hostname, sheet_name, False)

        if not is_nested:
            logging.debug("New Normal dict data...")

            if sheet_name.find("+") > 0:
                sheet_index = sheet_name.find("+")
                sheet_name = sheet_name[: sheet_index]
                logging.debug(sheet_name)

                if hostname in self.sheet_data.keys():
                    temp_host_data = self.sheet_data[hostname]
                    if sheet_name in temp_host_data.keys():
                        temp_sheet_data = temp_host_data[sheet_name]
                        temp_sheet_data.append(OrderedDict(dict_data))
                    else:
                        temp_sheet_data = OrderedDict()
                        temp_dict = [dict_data]
                        temp_sheet_data.update(
                            OrderedDict({sheet_name: temp_dict}))
                        temp_host_data.update(OrderedDict(temp_sheet_data))
                else:
                    temp_sheet_data = OrderedDict()
                    temp_dict = [dict_data]
                    temp_sheet_data.update(
                        OrderedDict({sheet_name: temp_dict}))
                    host_dict = OrderedDict()
                    host_dict.update(OrderedDict({hostname: temp_sheet_data}))
                    self.sheet_data.update(OrderedDict(host_dict))

            else:

                self.add_hierarical_data(
                    parentKey, dict_data, hostname, sheet_name, False)

    def parse_cell_value(self, value):

        if value == "None":
            value = None
        if type(value) is float:
            value = int(value)

        return value

    def process_by_sheet(self, wb, sheet_name):

        #sheet = wb.sheet_by_name(sheet_name)
        sheet = wb[sheet_name]
        if sheet_name == "base":
            logging.debug(
                "This is a base sheet.. Need to handle in different way..")

            hosts_header = 0
            key_header = 1
            value_header = 2

            logging.debug("Sheet has %d columns", sheet.max_column)
            for col in range(1,sheet.max_column+1):
                columnHeader = sheet.cell(1, col).value
                logging.debug("Col = %d, Header= %s",col,columnHeader)
                if columnHeader == "host":
                    hosts_header = col
                elif columnHeader == "name":
                    key_header = col
                elif columnHeader == "value":
                    value_header = col

            for row in range(1,sheet.max_row+1):

                if row != 1:
                    logging.debug("Row = %d, HostsHead = %d,keyhead = %d, valhead = %d",row,hosts_header,key_header,value_header)
                    hostname = sheet.cell(row, hosts_header).value
                    key = sheet.cell(row, key_header).value
                    value = self.parse_cell_value(
                        sheet.cell(row, value_header).value)
                    logging.debug("hostname = %s, key = %s, value=%s",hostname,key,value)
                    is_list = False
                    if key.find("+") > 0:
                        key_index = key.find("+")
                        key = key[: key_index]
                        is_list = True

                    if hostname in self.sheet_data.keys():
                        # temp_data=OrderedDict()
                        temp_data = self.sheet_data[hostname]
                        if is_list:
                            if key in temp_data.keys():
                                temp_data[key].append(value)
                            else:
                                temp_dict = {key: [value]}
                        else:
                            temp_dict = {key: value}
                        temp_data.update(OrderedDict(temp_dict))
                    else:
                        logging.debug("New sheet data..")
                        if is_list:
                            temp_dict = {key: [value]}
                        else:
                            temp_dict = {key: value}
                        host_dict = OrderedDict(
                            {hostname: OrderedDict(temp_dict)})
                        self.sheet_data.update(OrderedDict(host_dict))



        else:
            logging.debug("Other Sheets :: %s ", sheet_name)
            keys = []
            values = []
            keys_count = 0
            for row in range(1,sheet.max_row+1):
                is_header_row = False
                i = 0
                hostname = ""
                array_data = OrderedDict()
                empty_value_count = 0
                for col in range(1,sheet.max_column+1):
                    cell_value = sheet.cell(row, col).value
                    if cell_value is None:
                        cell_value = ""
                    if cell_value == "host":
                        is_header_row = True
                        keys = []

                    if cell_value == "---":
                        keys = []
                        break

                    if is_header_row:
                        keys.append(cell_value)
                    else:
                        values.append(cell_value)
                        if keys[i] == "host":
                            if cell_value != "":
                                hostname = cell_value
                            else:
                                empty_value_count = empty_value_count + 1
                        else:
                            logging.debug("Row = %i, col = %i, key = %s, cell_value = %s",row,col,keys[i],cell_value)
                            if keys[i] != "" and keys[i].find("+") > 0 or sheet_name.find("+") > 0:
                                temp_array_data = self.map_key_value(
                                    keys[i], cell_value, array_data)
                                array_data.update(temp_array_data)

                                if cell_value == "":
                                    empty_value_count = empty_value_count + 1

                            elif keys[i] != "":
                                self.process_data(
                                    keys[i], cell_value, hostname, sheet_name)
                        i = i + 1

                logging.debug(
                    "\n$$$$$$$$$$$$$$ array_data ::: %s ", array_data)
                if i == empty_value_count:
                    array_data = OrderedDict()

                if array_data:
                    if type(array_data) is OrderedDict:
                        parentKey = None
                        for key in pre_defined_unique_identifiers:
                            if "@"+key in array_data.keys():
                                parentKey = key
                                break
                        self.process_list_data(
                            parentKey, array_data, hostname, sheet_name)

        logging.debug(self.sheet_data)

    def convert_data(self):

        workbook = load_workbook(self.workbook_name)
        logging.debug(workbook.sheetnames)

        for sheet_name in workbook.sheetnames:

            logging.info("Processing sheet %s ", sheet_name)
            self.process_by_sheet(workbook, sheet_name)

            # Write content to yaml files.
            self.write_base_sheet_data(self.sheet_data, self.dest_dir)

        logging.debug("Spreadsheet converted successfully.")


def main():

    logging.debug(len(sys.argv))
    if len(sys.argv) != 3:
        logging.error(
            "\n\n Usage:\n\n  > xls2yaml.py <workbook name> <destination directory>\n\n Example:\n\n  > xls2yaml.py vDC.xlsx ./\n")
        #raise Exception("Arguments required: xls2yaml.py <workbook name> <destination directory>")
        return

    xls2yaml_instance = ExcelToYaml(sys.argv[1], sys.argv[2])
    xls2yaml_instance.convert_data()


if __name__ == '__main__':
    main()
