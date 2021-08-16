#!/usr/bin/env python3

""" ********************************************************

Project: nita-yaml-to-excel
Version: 21.7

Copyright (c) Juniper Networks, Inc., 2021. All rights reserved.

Notice and Disclaimer: This code is licensed to you under the Apache 2.0 License (the "License"). You may not use this code except in compliance with the License. This code is not an official Juniper product. You can obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0.html

SPDX-License-Identifier: Apache-2.0

Third-Party Code: This code may depend on other components under separate copyright notice and license terms. Your use of the source code for those components is subject to the terms and conditions of the respective license as noted in the Third-Party source code file.

******************************************************** """

import yaml
import sys
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Protection, Alignment
from openpyxl.utils import get_column_letter
from yaml.constructor import Constructor

import logging
logging.basicConfig(stream=sys.stderr, level=logging.INFO,
                    format='%(asctime)s: %(levelname)s: %(message)s')

style_count = 0


def add_bool(self, node):
    return self.construct_scalar(node)


Constructor.add_constructor(u'tag:yaml.org,2002:bool', add_bool)


def ordered_load(stream, Loader=yaml.Loader, object_pairs_hook=OrderedDict):
    class OrderedLoader(Loader):
        pass

    def construct_mapping(loader, node):
        loader.flatten_mapping(node)
        return object_pairs_hook(loader.construct_pairs(node))
    OrderedLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        construct_mapping)
    return yaml.load(stream, OrderedLoader)


class YamlToExcel (object):
    def __init__(self, *sysArgs):
        if type(sysArgs) is tuple:
            self.file_params = sysArgs[0]
        else:
            self.file_params = sysArgs

    # Adjust column width
    def column_auto_fit(self, ws, header_value, cell_value, column_index):
        max_column_width = 60
        min_column_width = 15

        column_header_width = len(header_value)
        column_value_width = len(str(cell_value))

        if str(cell_value).isupper():
            column_value_width = column_value_width + 5

        column_width = ws.column_dimensions[get_column_letter(
            column_index)].width

        # TODO this should never be None. Find out why
        if (column_width == None):
            column_width = 0

        if column_value_width > column_width:
            column_width = column_value_width + 5

        if column_header_width > column_width:
            column_width = column_header_width + 5

        if column_width > max_column_width:
            column_width = max_column_width

        if column_width < min_column_width:
            column_width = min_column_width

        ws.column_dimensions[get_column_letter(
            column_index)].width = column_width

    def add_column_header(self, ws, column_name, value, hostname, row_index):
        logging.debug("####################### %s %s %s ",
                      column_name, value, hostname)

        column_count = ws.max_column
        row_id = 1
        # Need to add column header in row = 1 & column = 1
        existing_column_id = 0
        for column_id in range(1, column_count + 1):
            cell_header_value = ws.cell(row=row_id, column=column_id).value
            logging.debug("$$ cell_header_value :: %s %s ",
                          cell_header_value, column_name)
            if cell_header_value == column_name or (column_name and cell_header_value and (cell_header_value == column_name[1:] or cell_header_value[1:] == column_name)):
                logging.debug(
                    "###Existing column header found in the sheet....")
                existing_column_id = column_id
                break
        logging.debug("Existing column id %s", existing_column_id)
        if existing_column_id == 0:
            ws.cell(row=row_id, column=1).value = "host"
            existing_column_id = ws.max_column + 1
            ws.cell(row=row_id, column=existing_column_id).value = column_name
            ws.cell(row=row_id, column=1).style = "header"
            ws.cell(row=row_id, column=existing_column_id).style = "header"

        row_column_index = {"row": row_index, "column": existing_column_id}

        return row_column_index

    def add_plain_dict_data(self, ws, repeatData, hostname, rIndex=1, addColumnHeader=True):
        logging.debug("RepeatData ::::::::::::::::::: %s ", repeatData)
        tempcolx = 2
        tempRIndex = ws.max_row
        if type(repeatData) is OrderedDict:
            for repkey in repeatData:

                row_column_index = self.add_column_header(ws, repkey, str(
                    repeatData[repkey]), hostname, tempRIndex)
                row_column_index = self.add_column_header(ws, repkey, str(
                    repeatData[repkey]), hostname, tempRIndex)

                column_index = 1
                if row_column_index != None:
                    column_index = row_column_index['column']
                    row_index = row_column_index['row']
                    if column_index == None:
                        column_index = ws.max_column + 1
                    if row_index:
                        if row_index > tempRIndex:
                            tempRIndex = row_index

                ws.cell(row=tempRIndex + 1, column=1).value = hostname
                ws.cell(row=tempRIndex + 1, column=column_index).value = self.parse_cell_value(repeatData[repkey])
                ws.cell(row = tempRIndex + 1, column = 1).style = "value"
                ws.cell(row = tempRIndex + 1, column = column_index).style = "value"

                # Adjust column width
                self.column_auto_fit(ws, "host", hostname, 1)
                self.column_auto_fit(ws, repkey, repeatData[repkey], column_index)

                tempcolx = tempcolx + 1
            rIndex = rIndex + 1
        return rIndex

    def build_dict_data(self, data, dictData=None, prefix="", put_unique_identifier=False):
        logging.debug("Building dict data ....%s %s %s ",
                      data, dictData, put_unique_identifier)
        if not dictData:
            dictData = OrderedDict()
        tDictData = OrderedDict(dictData)

        pre_defined_unique_identifiers = ['id', 'name', 'group']

        is_pk_exist = False
        unique_key = ""
        if put_unique_identifier:
            for pk in pre_defined_unique_identifiers:
                logging.debug("Unique key %s ", pk)
                if type(data) is OrderedDict:
                    if pk in data:
                        logging.debug(
                            "Pre defined unique identifier exist... %s", pk)
                        is_pk_exist = True
                        unique_key = pk
                        break

        if type(data) is OrderedDict:
            for key in data:
                logging.debug("Key %s ", key)
                logging.debug("Value %s", data[key])
                if is_pk_exist:
                    if unique_key == key:
                        is_pk_exist = False
                        put_unique_identifier = True
                    else:
                        put_unique_identifier = False

                if type(data[key]) is OrderedDict:
                    if prefix == "":
                        if put_unique_identifier == True:
                            temp_prefix = '@' + key
                            put_unique_identifier = False
                        else:
                            temp_prefix = key
                    else:
                        if put_unique_identifier == True:
                            if prefix.find('+') >= 0:
                                temp_prefix = prefix + '.' + '@' + key
                                put_unique_identifier = False
                            else:
                                temp_prefix = prefix + '.' + key
                                put_unique_identifier = False
                        else:
                            temp_prefix = prefix + '.' + key

                    tempDictData = self.build_dict_data(
                        data[key], tDictData, temp_prefix)
                    tDictData.update(tempDictData)
                elif type(data[key]) is not list:
                    if prefix == "":
                        if put_unique_identifier == True:
                            tDictData['@' + key] = data[key]
                            put_unique_identifier = False
                        else:
                            tDictData.update(OrderedDict({key: data[key]}))
                    else:
                        if put_unique_identifier == True:
                            if prefix.find('+') >= 0:
                                tempPre = prefix + '.' + '@' + key
                                put_unique_identifier = False
                            else:
                                tempPre = prefix + '.' + key
                                put_unique_identifier = False
                        else:
                            tempPre = prefix + '.' + key
                        tDictData.update(OrderedDict({tempPre: data[key]}))

        logging.debug("Final ::: tDictData >>> %s ", tDictData)
        return tDictData

    def build_list_data(self, data, listDataParam=None):
        logging.debug("Building list data.... %s %s ", data, listDataParam)
        if not listDataParam:
            listDataParam = OrderedDict()

        if type(data) is OrderedDict:
            for key in data:
                if (type(data[key]) is list):
                    logging.debug(
                        "-->>>>>>type(data[key]):::: %s", type(data[key]))
                    listDataParam[key] = data[key]

        return listDataParam

    def parse_recursive_data(self, ws, data, oldDictData=None, hostname="", add_column_header=True, rowIndex=1, colIndex=2, prefix=""):
        if not oldDictData:
            oldDictData = OrderedDict()

        tempOldDictData = OrderedDict(oldDictData)
        listData = self.build_list_data(data)

        logging.debug("listData ::::::: %s ", listData)

        if listData:
            logging.debug("\n\n\nFinal list in dictionary..... parse_recursive_data ...... %s %s %s %s ",
                          listData, rowIndex, tempOldDictData, prefix)
            tempPlainData = self.build_dict_data(
                data, oldDictData, prefix, True)
            logging.debug(
                "Dict data with unique identifier ::: parse_recursive_data %s ", tempPlainData)
            for listKey in listData:
                if prefix == "":
                    temp_prefix = listKey
                else:
                    temp_prefix = prefix + '.' + listKey + '+'
                rowIndex = self.parse_recursive_data(
                    ws, listData[listKey], tempPlainData, hostname, add_column_header, rowIndex, colIndex, temp_prefix)
                # rowIndex = rowIndex + 1

        if type(data) is list:
            logging.debug("$$$$$$$$$$$$$$$List.... %s ", data)
            tempRowIndex = rowIndex
            addColumnHeader = add_column_header
            for tempData in data:
                if type(tempData) is list:
                    logging.debug("To handle list ")
                elif type(tempData) is OrderedDict:
                    logging.debug(
                        "####################################### Dict .......... %s %s %s ", tempData, tempOldDictData, prefix)
                    temp_listData = self.build_list_data(tempData)
                    if not temp_listData:
                        tempDictData = self.build_dict_data(
                            tempData, tempOldDictData, prefix)
                        tempRowIndex = self.add_plain_dict_data(
                            ws, tempDictData, hostname, tempRowIndex, addColumnHeader)
                    else:
                        rowIndex = self.parse_recursive_data(
                            ws, tempData, oldDictData, hostname, add_column_header, rowIndex, colIndex, prefix)
                else:
                    logging.debug(
                        "Not dict and list :::::::: old data %s %s ", tempData, prefix)
                    tempRepeatData = tempOldDictData
                    tempRepeatData[prefix] = tempData
                    tempRowIndex = self.add_plain_dict_data(
                        ws, style, tempRepeatData, hostname, tempRowIndex, addColumnHeader)

            if tempRowIndex > rowIndex:
                rowIndex = tempRowIndex

        if type(data) is OrderedDict:
            for temp_key in data:
                if type(data[temp_key]) is OrderedDict:
                    temp_listData = self.build_list_data(data[temp_key])
                    if not temp_listData:
                        logging.debug("Not a list")
                    else:
                        if prefix == "":
                            temp_prefix = temp_key
                        else:
                            temp_prefix = prefix + '.' + temp_key + '+'
                        rowIndex = self.parse_recursive_data(
                            ws, style, data[temp_key], oldDictData, hostname, add_column_header, rowIndex, colIndex, temp_prefix)
                        logging.debug("rowIndex :: %s ", rowIndex)

        return rowIndex

    def generate_sheet_data(self, ws, data, oldDictData=None, hostname="", add_unique_field=False, add_column_header=True, rowIndex=1, colIndex=2, prefix=""):
        if not oldDictData:
            oldDictData = OrderedDict()

        tempOldDictData = OrderedDict(oldDictData)

        listData = self.build_list_data(data)
        tempPlainData = OrderedDict()

        if not listData:
            logging.debug("\n###########################List is empty...")
            tempDictData = self.build_dict_data(data, tempOldDictData, prefix)
            rowIndex = self.add_plain_dict_data(
                ws, tempDictData, hostname, rowIndex, add_column_header)
            logging.debug("testRowIndex :::::::::::: %s", rowIndex)
            for removeKey in tempDictData:
                data.pop(removeKey, None)
        else:
            logging.debug(
                "\n\n\nFinal list in dictionary..... %s %s %s ", listData, rowIndex, prefix)
            tempPlainData = self.build_dict_data(
                data, oldDictData, prefix, add_unique_field)
            logging.debug(
                "Dict data with unique identifier ::: %s", tempPlainData)
            for listKey in listData:
                logging.debug("List key :: %s ", listKey)
                if prefix == "":
                    temp_prefix = listKey + '+'
                else:
                    temp_prefix = prefix + '.' + listKey + '+'
                rowIndex = self.parse_recursive_data(
                    ws, listData[listKey], tempPlainData, hostname, add_column_header, rowIndex, colIndex, temp_prefix)

                if listKey in data:
                    del data[listKey]

            for removeKey in tempPlainData:
                if removeKey in data:
                    del data[removeKey]

        logging.debug("$$$$$$$$$$$$$$$$$$$$$$ data :::: %s", data)

        if data:
            rowIndex = self.parse_recursive_data(
                ws, data, tempPlainData, hostname, add_column_header, rowIndex, colIndex)

        logging.debug("$$$$$$$$$$$$$$$$$$$$$$ rowIndex :::: %s", rowIndex)

        return rowIndex

    def parse_yaml_files(self, wb, ws, file_content, hostname, sheet_last_row_index):
        for key in file_content:
            keyValuePair = []
            keyValuePair.append(key)
            value = file_content[key]

            if type(value) is OrderedDict:
                is_sheet_already_exist = False
                try:
                    wb[key]
                    is_sheet_already_exist = True
                except:
                    logging.debug("Sheet doesn't exist. Creating new sheet.")

                temp_row_index = 1
                add_column_header = True
                if is_sheet_already_exist:
                    additionalws = wb[key]
                    existing_sheet_row_index = sheet_last_row_index[key]
                    temp_row_index = existing_sheet_row_index
                else:
                    additionalws = wb.create_sheet(title=key)
                    additionalws.cell(row=1, column=1).value = "host"

                tempValue = value
                additional_sheet_row_index = self.generate_sheet_data(
                    additionalws, value, OrderedDict(), hostname, False, add_column_header, temp_row_index)
                sheet_last_row_index[key] = additional_sheet_row_index

            elif type(value) is list:

                logging.debug(
                    "\nIt is a list -- Key :: %s Value :: %s ", key, value)
                key = key + "+"
                is_sheet_already_exist = False
                try:
                    wb[key]
                    is_sheet_already_exist = True
                except:
                    logging.debug("Sheet doesn't exist. Creating new sheet.")

                tempRowIndex = 1
                add_column_header = True
                if is_sheet_already_exist:
                    additionalws = wb[key]
                    existing_sheet_row_index = sheet_last_row_index[key]
                    tempRowIndex = existing_sheet_row_index
                    # add_column_header = False
                else:
                    additionalws = wb.create_sheet(title=key)

                need_remove_sheet = False
                for listTemp in value:
                    logging.debug("\n listTemp ::: %s ", listTemp)
                    if type(listTemp) is list or type(listTemp) is OrderedDict:
                        testRowIndex = self.generate_sheet_data(
                            additionalws, listTemp, OrderedDict(), hostname, True, False, tempRowIndex)
                        tempRowIndex = tempRowIndex + 1
                    else:
                        logging.debug("Not a list or dict...")
                        base_sheet_rowx = sheet_last_row_index['base']
                        ws.cell(row = base_sheet_rowx + 1, column = 1).value = hostname
                        ws.cell(row = base_sheet_rowx + 1, column = 1).style = "value"
                        ws.cell(row = base_sheet_rowx + 1, column = 2).value = key
                        ws.cell(row = base_sheet_rowx + 1, column = 2).style = "value"
                        ws.cell(row = base_sheet_rowx + 1, column = 3).value = self.parse_cell_value(listTemp)
                        ws.cell(row = base_sheet_rowx + 1, column = 3).style = "value"

                        sheet_last_row_index['base'] = base_sheet_rowx + 1
                        need_remove_sheet = True

                # Clean up sheet
                if need_remove_sheet:
                    cleanupsheet = wb[key]
                    if cleanupsheet is not None:
                        wb.remove(cleanupsheet)

                sheet_last_row_index[key] = tempRowIndex
            else:
                logging.debug("base:" + key + "\n")
                logging.debug("type:" + str(type(value)) + "\n")
                logging.debug(self.parse_cell_value(file_content[key]))

                base_sheet_rowx = sheet_last_row_index['base']
                ws.cell(row=base_sheet_rowx + 1, column=1).value = hostname
                ws.cell(row=base_sheet_rowx + 1, column=1).style = "value"
                ws.cell(row=base_sheet_rowx + 1, column=2).value = key
                ws.cell(row=base_sheet_rowx + 1, column=2).style = "value"
                ws.cell(row=base_sheet_rowx + 1, column=3).value = self.parse_cell_value(file_content[key])
                ws.cell(row=base_sheet_rowx + 1, column=3).style = "value"

                # Adjust column width
                self.column_auto_fit(ws, "host", hostname, 1)
                self.column_auto_fit(ws, "Name", key, 2)
                self.column_auto_fit(ws, "Value", file_content[key], 3)

                sheet_last_row_index['base'] = base_sheet_rowx + 1

    def parse_cell_value(self, value):
        if type(value) is bool:
            value = str(value)
        elif type(value) is int:
            value = value
        else:
            value = str(value)

        return value

    def put_border(self, wb):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for sheet in wb:
            for rows in sheet.rows:
                for cell in rows:
                    cell.border = border

    def convert_data(self):
        wb = Workbook()

        ws = wb.active
        ws.title = 'base'

        value_font = Font(name="Bitstream Charter", size=10)
        vthin = Side(border_style="thin", color="000000")
        vborder = Border(top=vthin, left=vthin, right=vthin, bottom=vthin)
        valignment = Alignment(wrap_text=True)

        value_style = NamedStyle(name = "value", font = value_font, border = vborder, alignment = valignment)
        wb.add_named_style(value_style)

        header_font = Font(name="Bitstream Charter", size=10, bold=True)
        hthin = Side(border_style="thin", color="000000")
        hfill = PatternFill(fill_type='solid', fgColor="33bbff")
        hborder = Border(top=hthin, left=hthin, right=hthin, bottom=hthin)

        header_style = NamedStyle(name = "header", font = header_font, fill = hfill, border = hborder)
        wb.add_named_style(header_style)

        ws.cell(row=1, column=1).value = "host"
        ws.cell(row=1, column=1).style = "header"
        ws.cell(row=1, column=2).value = "name"
        ws.cell(row=1, column=2).style = "header"
        ws.cell(row=1, column=3).value = "value"
        ws.cell(row=1, column=3).style = "header"

        sheet_last_row_index = {}

        sheet_last_row_index['base'] = 1

        logging.debug(self.file_params)

        param_len = len(self.file_params)
        if param_len == 1:
            if not self.file_params[0].endswith(".yaml") and not self.file_params[0].endswith(".yml"):
                logging.error(
                    "\n\n Usage:\n\n > yaml2xls.py <yaml-files-path> [spreadsheet-name] Optional \n\n Examples:\n\n > yaml2xls.py group_vars/* host_vars/* vDC.xlsx\n > yaml2xls.py group_vars/* vDC.xlsx\n")
                return

        spreadsheet_file_name = ""
        count = 1
        yaml_file_count = 0
        for file_name in self.file_params:
            if not file_name.endswith("yaml2xls.py"):
                if count == param_len:
                    if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                        spreadsheet_file_name = file_name
                    else:
                        if not file_name.endswith(".yaml") and not file_name.endswith(".yml"):
                            logging.error(
                                "Invalid spreadsheet file extension given.")
                            return
                        else:
                            yaml_file_count = yaml_file_count + 1
                else:
                    if not file_name.endswith(".yaml") and not file_name.endswith(".yml"):
                        logging.error(
                            "'%s' file is an invalid file. Please give .yaml or .yml file.", file_name)
                        return
                    else:
                        yaml_file_count = yaml_file_count + 1

            count = count + 1

        if yaml_file_count == 0:
            logging.error(
                'Please give at least one .yaml or .yml file as input.')
            return

        isEmpty = True
        for file_name in self.file_params:
            if file_name.endswith(".yaml") or file_name.endswith(".yml"):
                logging.debug("YAML or YML file : " + file_name)

                with open(file_name, 'r') as stream:
                    content = ordered_load(stream, yaml.SafeLoader)

                host_name = file_name
                self.parse_yaml_files(
                    wb, ws, content, host_name, sheet_last_row_index)
                isEmpty = False

        if not isEmpty:
            #self.put_border(wb)
            logging.debug("YAML to excel conversion completed.")
            if spreadsheet_file_name:
                wb.save(spreadsheet_file_name)
            else:
                wb.save("all.xlsx")

if __name__ == '__main__':
    YamlToExcel(sys.argv).convert_data()
